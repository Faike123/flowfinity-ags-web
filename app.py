from __future__ import annotations

import html
import csv
import io
import json
import re
import zipfile
from datetime import datetime
import pandas as pd
import streamlit as st
from openpyxl import Workbook

from flowfinity_ags.engine import DEFAULT_SETTINGS, run_from_zip_bytes
from flowfinity_ags.gas_engine import run_gas_from_uploaded_sources
from flowfinity_ags.csv_ags_engine import run_csv_ags_from_uploaded_sources
from flowfinity_ags.profiles import AGS_PROFILE_OPTIONS


# ============================================================
# PAGE SETUP
# ============================================================

st.set_page_config(
    page_title="AGS Transformer",
    page_icon="🟨",
    layout="wide",
)


# ============================================================
# STYLE
# ============================================================

st.markdown(
    """
    <style>
        :root {
            --igne-green-dark: #003F23;
            --igne-green: #006B3F;
            --igne-green-2: #0A5A37;
            --igne-green-pale: #EAF2E9;
            --igne-green-soft: #F3F7F1;
            --igne-gold: #FBB800;
            --igne-gold-dark: #D99A00;
            --igne-cream: #F8F5EA;
            --igne-card: #FFFFFF;
            --igne-border: #DCE4D8;
            --igne-text: #13261B;
            --igne-muted: #5B675F;
        }

        .stApp {
            background:
                radial-gradient(circle at top right, rgba(251,184,0,0.10), transparent 32%),
                linear-gradient(180deg, #F8F5EA 0%, #F3F7F1 44%, #FFFFFF 100%);
            color: var(--igne-text);
        }

        h1, h2, h3 {
            color: var(--igne-green-dark) !important;
            letter-spacing: -0.025em;
        }

        [data-testid="stSidebar"] {
            background:
                linear-gradient(180deg, #003F23 0%, #004E2E 55%, #002B18 100%);
            border-right: 6px solid var(--igne-gold);
        }

        [data-testid="stSidebar"] h1,
        [data-testid="stSidebar"] h2,
        [data-testid="stSidebar"] h3,
        [data-testid="stSidebar"] label,
        [data-testid="stSidebar"] p,
        [data-testid="stSidebar"] span {
            color: #FFFFFF !important;
        }

        [data-testid="stSidebar"] input,
        [data-testid="stSidebar"] textarea {
            background: #FFFDF4 !important;
            color: #13261B !important;
            border: 1px solid rgba(251, 184, 0, 0.70) !important;
            border-radius: 8px !important;
        }

        [data-testid="stSidebar"] div[data-baseweb="select"] > div {
            background: #FFFDF4 !important;
            border: 1px solid rgba(251, 184, 0, 0.70) !important;
            border-radius: 8px !important;
        }

        [data-testid="stSidebar"] div[data-baseweb="select"] span {
            color: #13261B !important;
        }

        [data-testid="stSidebar"] hr {
            border-color: rgba(251,184,0,0.35);
        }

        .brand-block {
            margin: 0 0 1.6rem 0;
            padding: 0.2rem 0 0.5rem 0;
        }

        .brand-logo {
            font-size: 2.45rem;
            font-weight: 900;
            letter-spacing: 0.02em;
            color: white;
            line-height: 1;
        }

        .brand-logo span {
            color: var(--igne-gold);
        }

        .brand-subtitle {
            color: rgba(255,255,255,0.72);
            font-size: 0.74rem;
            letter-spacing: 0.18em;
            text-transform: uppercase;
            margin-top: 0.35rem;
        }

        .sidebar-note {
            margin-top: 2rem;
            padding-top: 1rem;
            border-top: 1px solid rgba(255,255,255,0.18);
            color: rgba(255,255,255,0.72);
            font-size: 0.82rem;
        }

        .hero-row {
            display: flex;
            align-items: center;
            gap: 1.2rem;
            margin: 1.2rem 0 1.4rem 0;
        }

        .hero-icon {
            width: 76px;
            height: 76px;
            border-radius: 999px;
            background: var(--igne-green-pale);
            border: 1px solid var(--igne-border);
            display: flex;
            align-items: center;
            justify-content: center;
            color: var(--igne-green-dark);
            font-size: 2.2rem;
            box-shadow: 0 8px 20px rgba(0,63,35,0.08);
        }

        .hero-title {
            font-size: 2.55rem;
            line-height: 1;
            font-weight: 900;
            color: var(--igne-green-dark);
            margin: 0;
        }

        .hero-subtitle {
            color: var(--igne-muted);
            margin-top: 0.45rem;
            font-size: 1rem;
        }

        .upload-card {
            background: rgba(255,255,255,0.88);
            border: 1px solid var(--igne-border);
            border-radius: 16px;
            padding: 1.35rem 1.55rem;
            margin-bottom: 1rem;
            box-shadow: 0 8px 24px rgba(0,63,35,0.06);
        }

        .upload-title {
            color: var(--igne-green-dark);
            font-weight: 850;
            font-size: 1.08rem;
            margin-bottom: 0.8rem;
        }

        div[data-testid="stFileUploader"] section {
            background: #FFFFFF !important;
            border: 2px dashed rgba(0,63,35,0.55) !important;
            border-radius: 14px !important;
            padding: 1.1rem !important;
        }

        div[data-testid="stFileUploader"] section:hover {
            border-color: var(--igne-gold) !important;
            background: #FFFBEA !important;
        }

        div[data-testid="stFileUploader"] button {
            background: var(--igne-green-dark) !important;
            color: #FFFFFF !important;
            border: 1px solid var(--igne-green-dark) !important;
            border-radius: 9px !important;
            font-weight: 800 !important;
        }

        .tip-box {
            background: #F7FAF4;
            border: 1px solid var(--igne-border);
            border-left: 5px solid var(--igne-gold);
            border-radius: 12px;
            padding: 0.9rem 1rem;
            color: var(--igne-text);
        }

        .tip-box strong {
            color: var(--igne-green-dark);
        }

        .summary-grid {
            display: grid;
            grid-template-columns: repeat(4, minmax(0, 1fr));
            gap: 1rem;
            margin: 1rem 0 1.2rem 0;
        }

        .summary-card {
            background: var(--igne-card);
            border: 1px solid var(--igne-border);
            border-left: 5px solid var(--igne-gold);
            border-radius: 14px;
            padding: 1rem 1.05rem;
            box-shadow: 0 7px 18px rgba(0,63,35,0.06);
            display: flex;
            gap: 0.9rem;
            align-items: center;
            min-height: 98px;
        }

        .summary-icon {
            width: 52px;
            height: 52px;
            border-radius: 999px;
            background: var(--igne-green-pale);
            display: flex;
            align-items: center;
            justify-content: center;
            color: var(--igne-green-dark);
            font-size: 1.55rem;
            flex-shrink: 0;
        }

        .summary-label {
            color: var(--igne-muted);
            font-size: 0.88rem;
            font-weight: 750;
        }

        .summary-value {
            color: var(--igne-green-dark);
            font-size: 1.75rem;
            font-weight: 900;
            line-height: 1.05;
        }

        .summary-caption {
            color: var(--igne-muted);
            font-size: 0.82rem;
        }

        .section-card {
            background: rgba(255,255,255,0.9);
            border: 1px solid var(--igne-border);
            border-radius: 15px;
            padding: 1rem 1.15rem;
            margin-bottom: 1rem;
            box-shadow: 0 7px 18px rgba(0,63,35,0.05);
        }

        .section-heading {
            color: var(--igne-green-dark);
            font-weight: 900;
            font-size: 1.08rem;
            margin-bottom: 0.8rem;
        }

        .chip-row {
            display: flex;
            flex-wrap: wrap;
            gap: 0.7rem;
        }

        .group-chip {
            border: 1px solid var(--igne-border);
            background: linear-gradient(180deg, #FFFFFF 0%, #F7FAF5 100%);
            color: var(--igne-green-dark);
            border-radius: 10px;
            padding: 0.55rem 0.85rem;
            font-weight: 850;
            min-width: 98px;
            text-align: center;
            box-shadow: 0 3px 10px rgba(0,63,35,0.05);
        }

        .status-pill {
            display: inline-flex;
            align-items: center;
            gap: 0.35rem;
            background: #DFF2E1;
            color: var(--igne-green-dark);
            border-radius: 999px;
            padding: 0.25rem 0.65rem;
            font-weight: 850;
            font-size: 0.82rem;
        }

        div[data-testid="stDataFrame"] {
            border: 1px solid var(--igne-border);
            border-radius: 12px;
            overflow: hidden;
            box-shadow: 0 5px 14px rgba(0,63,35,0.04);
        }

        [data-testid="stTabs"] button {
            font-weight: 800;
        }

        [data-testid="stTabs"] button[aria-selected="true"] {
            color: var(--igne-green-dark) !important;
            border-bottom: 4px solid var(--igne-gold) !important;
        }

        .stButton > button {
            background: var(--igne-green-dark) !important;
            color: white !important;
            border: 1px solid var(--igne-green-dark) !important;
            border-radius: 10px !important;
            font-weight: 850 !important;
        }

        .stButton > button:hover {
            background: var(--igne-green) !important;
            border-color: var(--igne-green) !important;
            color: white !important;
        }

        .stDownloadButton > button {
            background: var(--igne-gold) !important;
            color: #13261B !important;
            border: 1px solid var(--igne-gold-dark) !important;
            border-radius: 10px !important;
            font-weight: 900 !important;
        }

        .stDownloadButton > button:hover {
            background: #E5A900 !important;
            color: #13261B !important;
            border-color: #C89100 !important;
        }

        .small-muted {
            color: var(--igne-muted);
            font-size: 0.88rem;
        }

        @media (max-width: 1100px) {
            .summary-grid {
                grid-template-columns: repeat(2, minmax(0, 1fr));
            }
        }
    </style>
    """,
    unsafe_allow_html=True,
)



st.markdown(
    """
    <style>
        /* FINAL READABILITY OVERRIDES */

        /* Main app text should never disappear into the pale background */
        .main *, 
        [data-testid="stAppViewContainer"] .main *,
        [data-testid="stVerticalBlock"] *,
        [data-testid="stHorizontalBlock"] * {
            color: #13261B;
        }

        /* Keep sidebar text white, but inputs dark */
        [data-testid="stSidebar"] h1,
        [data-testid="stSidebar"] h2,
        [data-testid="stSidebar"] h3,
        [data-testid="stSidebar"] label,
        [data-testid="stSidebar"] p,
        [data-testid="stSidebar"] span {
            color: #FFFFFF !important;
        }

        [data-testid="stSidebar"] input,
        [data-testid="stSidebar"] textarea,
        [data-testid="stSidebar"] input *,
        [data-testid="stSidebar"] textarea * {
            color: #13261B !important;
            background: #FFFDF4 !important;
        }

        [data-testid="stSidebar"] div[data-baseweb="select"] span,
        [data-testid="stSidebar"] div[data-baseweb="select"] svg {
            color: #13261B !important;
            fill: #13261B !important;
        }

        /* Selectbox in main area */
        div[data-baseweb="select"] span {
            color: #13261B !important;
        }

        /* Metrics */
        div[data-testid="stMetric"],
        div[data-testid="stMetric"] * {
            color: #13261B !important;
        }

        div[data-testid="stMetricValue"] {
            color: #003F23 !important;
        }

        div[data-testid="stMetricLabel"] p {
            color: #4A5A50 !important;
        }

        /* Tabs */
        [data-testid="stTabs"] button,
        [data-testid="stTabs"] button p,
        [data-testid="stTabs"] button span {
            color: #44524A !important;
            font-weight: 800 !important;
        }

        [data-testid="stTabs"] button[aria-selected="true"],
        [data-testid="stTabs"] button[aria-selected="true"] p,
        [data-testid="stTabs"] button[aria-selected="true"] span {
            color: #003F23 !important;
            font-weight: 900 !important;
        }

        /* Dataframes */
        div[data-testid="stDataFrame"] {
            background: #FFFFFF !important;
        }

        div[data-testid="stDataFrame"] * {
            color: #13261B !important;
        }

        /* Expander */
        details,
        details * {
            color: #13261B !important;
        }

        /* File uploader button */
        div[data-testid="stFileUploader"] button,
        div[data-testid="stFileUploader"] button * {
            background: #003F23 !important;
            color: #FFFFFF !important;
            border-color: #003F23 !important;
        }

        /* Primary buttons */
        .stButton > button,
        .stButton > button * {
            background: #003F23 !important;
            color: #FFFFFF !important;
            border-color: #003F23 !important;
        }

        .stButton > button:hover,
        .stButton > button:hover * {
            background: #006B3F !important;
            color: #FFFFFF !important;
            border-color: #006B3F !important;
        }

        /* Download/export button */
        .stDownloadButton > button,
        .stDownloadButton > button * {
            background: #FBB800 !important;
            color: #13261B !important;
            border-color: #D99A00 !important;
        }

        /* Alerts */
        [data-testid="stAlert"],
        [data-testid="stAlert"] * {
            color: #13261B !important;
        }

        /* Remove awkward blank white pill near title if any empty Streamlit block gets styled */
        .element-container:empty {
            display: none !important;
        }
    </style>
    """,
    unsafe_allow_html=True,
)


# ============================================================
# SIDEBAR
# ============================================================

with st.sidebar:
    st.markdown(
        """
        <div class="brand-block">
            <div class="brand-logo"><span>✦</span> AGS Transformer</div>
            <div class="brand-subtitle">Field data converter</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.header("1. Export settings")

    transformer_options = [
        "Flowfinity field data to AGS",
        "Field gas monitoring to AGS",
        "CSV / Excel AGS tables to AGS",
    ]

    transformer_id_by_name = {
        "Flowfinity field data to AGS": "ff_to_ags",
        "Field gas monitoring to AGS": "gas_to_ags",
        "CSV / Excel AGS tables to AGS": "csv_to_ags",
    }

    transformer_name = st.selectbox(
        "Transformer",
        transformer_options,
        index=0,
        help="Choose which source format to transform into AGS.",
    )

    selected_transformer = transformer_id_by_name[transformer_name]

    st.session_state["selected_transformer"] = selected_transformer

    if st.session_state.get("_last_selected_transformer") != selected_transformer:
        st.session_state["result"] = None
        st.session_state["uploaded_source_files"] = []
        st.session_state["uploaded_zip_bytes"] = None
        st.session_state["selected_project_id"] = None

    st.session_state["_last_selected_transformer"] = selected_transformer

    st.caption(
        "The same upload, preview, and export layout will be reused for each transformer."
    )

    st.divider()

    profile_name = st.selectbox(
        "AGS format",
        AGS_PROFILE_OPTIONS,
        index=0,
    )

    st.caption(
        "Current profiles are compact Flowfinity export profiles. "
        "Full official AGS heading profiles can be added later."
    )

    st.divider()

    st.header("2. Project metadata")
    st.caption("Optional. Leave blank to use project values detected from the uploaded source data.")

    sidebar_project_id = st.text_input("Project ID override", value="", help="Optional. Exports to AGS field PROJ_ID")
    sidebar_project_name = st.text_input("Project name", value="", help="Optional. Exports to AGS field PROJ_NAME")
    sidebar_project_location = st.text_input("Project location", value="", help="Optional. Exports to AGS field PROJ_LOC")
    sidebar_project_client = st.text_input("Client", value="", help="Optional. Exports to AGS field PROJ_CLNT")
    sidebar_project_contractor = st.text_input("Contractor", value="", help="Optional. Exports to AGS field PROJ_CONT")
    sidebar_project_engineer = st.text_input("Engineer", value="", help="Optional. Exports to AGS field PROJ_ENG")

    st.divider()

    st.header("3. Transfer metadata")

    tran_prod = st.text_input("Produced by", value=DEFAULT_SETTINGS["tran_prod"], help="Exports to AGS field TRAN_PROD")
    tran_stat = st.text_input("Transfer status", value=DEFAULT_SETTINGS["tran_stat"], help="Exports to AGS field TRAN_STAT")
    tran_desc = st.text_input("Transfer description", value=("Field gas monitoring export" if selected_transformer == "gas_to_ags" else ("CSV / Excel AGS table import" if selected_transformer == "csv_to_ags" else DEFAULT_SETTINGS["tran_desc"])), help="Exports to AGS field TRAN_DESC")
    tran_recv = st.text_input("Received by", value=DEFAULT_SETTINGS["tran_recv"], help="Exports to AGS field TRAN_RECV")
    tran_dlim = st.text_input("Data delimiter", value=DEFAULT_SETTINGS["tran_dlim"], help="Exports to AGS field TRAN_DLIM")
    tran_rcon = st.text_input("Continuation symbol", value=DEFAULT_SETTINGS["tran_rcon"], help="Exports to AGS field TRAN_RCON")

    st.divider()

    st.header("4. Location defaults")

    default_loca_type = st.text_input("Default location type", value=("BH" if selected_transformer == "gas_to_ags" else DEFAULT_SETTINGS["default_loca_type"]), help="Exports to AGS field LOCA_TYPE")
    default_loca_rem = st.text_input("Default location remark", value=DEFAULT_SETTINGS["default_loca_rem"], help="Exports to AGS field LOCA_REM")

    st.markdown(
        """
        <div class="sidebar-note">
            🛡️ Test prototype. Use development data only.
        </div>
        """,
        unsafe_allow_html=True,
    )


base_settings = {
    "project_id": sidebar_project_id.strip(),
    "project_name": sidebar_project_name.strip(),
    "project_location": sidebar_project_location.strip(),
    "project_client": sidebar_project_client.strip(),
    "project_contractor": sidebar_project_contractor.strip(),
    "project_engineer": sidebar_project_engineer.strip(),
    "project_memo": "",
    "tran_prod": tran_prod,
    "tran_stat": tran_stat,
    "tran_desc": tran_desc,
    "tran_recv": tran_recv,
    "tran_dlim": tran_dlim,
    "tran_rcon": tran_rcon,
    "default_loca_type": default_loca_type,
    "default_loca_rem": default_loca_rem,
}






active_transformer = st.session_state.get("selected_transformer", "ff_to_ags")

if active_transformer == "gas_to_ags":
    transformer_title = "Field gas monitoring to AGS"
    transformer_subtitle = (
        "Upload environmental gas monitoring data, inspect mapped readings, "
        "and export AGS using the same workflow."
    )
    hero_icon = "◌"
    upload_title = "1. Upload field gas monitoring data"
    upload_label = "Upload gas monitoring Excel or ZIP"
    upload_help = "Use an Excel gas monitoring sheet or a ZIP containing multiple gas monitoring sheets."
    accepted_upload_types = ["zip", "xlsx", "xlsm", "xls"]
    process_button_label = "Process gas monitoring data"
elif active_transformer == "csv_to_ags":
    transformer_title = "CSV / Excel AGS tables to AGS"
    transformer_subtitle = (
        "Upload CSV files, Excel workbooks, or ZIP batches where each file or sheet "
        "represents an AGS group, then review and export one AGS file."
    )
    hero_icon = "▦"
    upload_title = "1. Upload CSV / Excel AGS tables"
    upload_label = "Upload CSV, Excel, or ZIP batch"
    upload_help = "Use CSV files, Excel workbooks, multiple files, or a ZIP containing CSV / Excel AGS tables."
    accepted_upload_types = ["zip", "csv", "xlsx", "xlsm"]
    process_button_label = "Process CSV / Excel AGS tables"
else:
    transformer_title = "Flowfinity field data to AGS"
    transformer_subtitle = (
        "Upload Flowfinity field data ZIP, select project, inspect datasets and locations, "
        "then export AGS."
    )
    hero_icon = "⇧"
    upload_title = "1. Upload Flowfinity field data"
    upload_label = "Upload Flowfinity ZIP"
    upload_help = "Use the ZIP export produced by the current Flowfinity workflow."
    accepted_upload_types = ["zip"]
    process_button_label = "Process field data"




def parse_ags_text_to_group_tables(ags_text: str) -> dict[str, dict]:
    """
    Convert AGS text into table payloads so any transformer output can be
    previewed/exported as normal CSV and JSON tables.
    """
    groups: dict[str, dict] = {}
    current_group = ""

    reader = csv.reader(io.StringIO(ags_text))

    for row in reader:
        if not row:
            continue

        control = str(row[0]).strip().upper()

        if control == "GROUP":
            current_group = str(row[1]).strip() if len(row) > 1 else ""

            if current_group:
                groups[current_group] = {
                    "headings": [],
                    "units": [],
                    "types": [],
                    "rows": [],
                }

            continue

        if not current_group or current_group not in groups:
            continue

        payload = groups[current_group]

        if control == "HEADING":
            payload["headings"] = [str(value) for value in row[1:]]
        elif control == "UNIT":
            payload["units"] = [str(value) for value in row[1:]]
        elif control == "TYPE":
            payload["types"] = [str(value) for value in row[1:]]
        elif control == "DATA":
            headings = payload.get("headings", [])
            values = [str(value) for value in row[1:]]

            if headings:
                values = pad_export_row(values, len(headings))
                payload["rows"].append(values)

    return groups


def pad_export_row(values: list[str], length: int) -> list[str]:
    output = list(values)

    while len(output) < length:
        output.append("")

    return output[:length]


def unique_table_fields(headings: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    fields: list[str] = []

    for index, heading in enumerate(headings):
        base = str(heading).strip() or f"FIELD_{index + 1}"
        count = seen.get(base, 0) + 1
        seen[base] = count

        if count == 1:
            fields.append(base)
        else:
            fields.append(f"{base}_{count}")

    return fields


def ags_group_to_records(table: dict) -> list[dict[str, str]]:
    headings = table.get("headings", [])
    fields = unique_table_fields(headings)
    records: list[dict[str, str]] = []

    for row in table.get("rows", []):
        padded_row = pad_export_row(row, len(fields))
        records.append({fields[index]: padded_row[index] for index in range(len(fields))})

    return records


def ags_group_to_json_payload(group_name: str, table: dict) -> dict:
    headings = table.get("headings", [])
    units = pad_export_row(table.get("units", []), len(headings))
    types = pad_export_row(table.get("types", []), len(headings))

    return {
        "group": group_name,
        "headings": headings,
        "units": units,
        "types": types,
        "row_count": len(table.get("rows", [])),
        "rows": ags_group_to_records(table),
    }


def ags_group_to_csv_bytes(table: dict) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\n")

    headings = table.get("headings", [])
    writer.writerow(headings)

    for row in table.get("rows", []):
        writer.writerow(pad_export_row(row, len(headings)))

    return output.getvalue().encode("utf-8-sig")


def ags_group_to_json_bytes(group_name: str, table: dict) -> bytes:
    payload = ags_group_to_json_payload(group_name, table)
    return json.dumps(payload, indent=2, ensure_ascii=False).encode("utf-8")


def safe_export_name(value: str, fallback: str = "ags_export") -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", str(value or "").strip()).strip("_")

    if not cleaned:
        return fallback

    return cleaned[:80]



def safe_excel_sheet_name(value: str, used_names: set[str]) -> str:
    cleaned = re.sub(r"[\[\]\:\*\?\/\\]+", "_", str(value or "").strip())
    cleaned = re.sub(r"\s+", "_", cleaned).strip("_")

    if not cleaned:
        cleaned = "GROUP"

    base = cleaned[:31]
    candidate = base
    suffix = 2

    while candidate in used_names:
        suffix_text = f"_{suffix}"
        candidate = f"{base[:31 - len(suffix_text)]}{suffix_text}"
        suffix += 1

    used_names.add(candidate)
    return candidate


def build_excel_workbook_bytes(ags_text: str, base_name: str = "ags_tables") -> bytes:
    groups = parse_ags_text_to_group_tables(ags_text)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    manifest_sheet = workbook.create_sheet("manifest")
    manifest_sheet.append(["package_name", safe_export_name(base_name)])
    manifest_sheet.append(["group_count", len(groups)])
    manifest_sheet.append(["created_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    manifest_sheet.append([])
    manifest_sheet.append(["group", "sheet_name", "rows", "fields"])

    used_sheet_names = {"manifest"}

    for group_name, table in groups.items():
        sheet_name = safe_excel_sheet_name(group_name, used_sheet_names)
        headings = table.get("headings", [])
        rows = table.get("rows", [])

        manifest_sheet.append([group_name, sheet_name, len(rows), len(headings)])

        worksheet = workbook.create_sheet(sheet_name)
        worksheet.append(headings)

        for row in rows:
            worksheet.append(pad_export_row(row, len(headings)))

        worksheet.freeze_panes = "A2"

        if headings:
            worksheet.auto_filter.ref = worksheet.dimensions

        for column_cells in worksheet.columns:
            column_letter = column_cells[0].column_letter
            max_length = 8

            for cell in column_cells[:200]:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, min(len(value), 40))

            worksheet.column_dimensions[column_letter].width = max_length + 2

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_table_package_zip(
    ags_text: str,
    base_name: str,
    include_ags: bool = False,
    include_csv: bool = False,
    include_json: bool = False,
    include_excel: bool = False,
) -> bytes:
    groups = parse_ags_text_to_group_tables(ags_text)
    zip_buffer = io.BytesIO()
    safe_base = safe_export_name(base_name)

    with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        manifest = {
            "package_name": safe_base,
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "group_count": len(groups),
            "groups": [
                {
                    "group": group_name,
                    "row_count": len(table.get("rows", [])),
                    "heading_count": len(table.get("headings", [])),
                }
                for group_name, table in groups.items()
            ],
            "contains": {
                "ags": include_ags,
                "csv": include_csv,
                "json": include_json,
                "excel": include_excel,
            },
        }

        archive.writestr("manifest.json", json.dumps(manifest, indent=2, ensure_ascii=False))

        if include_ags:
            archive.writestr(f"{safe_base}.ags", ags_text.encode("utf-8"))

        if include_excel:
            archive.writestr(
                f"{safe_base}.xlsx",
                build_excel_workbook_bytes(ags_text, safe_base),
            )

        for group_name, table in groups.items():
            safe_group = safe_export_name(group_name, "GROUP")

            if include_csv:
                archive.writestr(f"csv/{safe_group}.csv", ags_group_to_csv_bytes(table))

            if include_json:
                archive.writestr(f"json/{safe_group}.json", ags_group_to_json_bytes(group_name, table))

    return zip_buffer.getvalue()


def extract_ags_group_text(ags_text: str, group_name: str) -> str:
    """
    Extract one AGS group block from exported AGS text.
    """
    lines = ags_text.splitlines()
    output_lines = []
    in_group = False

    group_marker = f'"GROUP","{group_name}"'

    for line in lines:
        if line.startswith('"GROUP",'):
            if in_group:
                break
            if line.strip() == group_marker:
                in_group = True

        if in_group:
            output_lines.append(line)

    return "\n".join(output_lines).strip()


def ags_group_to_dataframe(group_text: str) -> pd.DataFrame:
    """
    Convert one AGS group block into a simple dataframe of DATA rows.
    """
    if not group_text.strip():
        return pd.DataFrame()

    rows = list(csv.reader(io.StringIO(group_text)))
    headings = []
    data_rows = []

    for row in rows:
        if not row:
            continue
        if row[0] == "HEADING":
            headings = row[1:]
        elif row[0] == "DATA":
            data_rows.append(row[1:])

    if not headings:
        return pd.DataFrame()

    normalised_rows = []
    for row in data_rows:
        if len(row) < len(headings):
            row = row + [""] * (len(headings) - len(row))
        elif len(row) > len(headings):
            row = row[:len(headings)]
        normalised_rows.append(row)

    return pd.DataFrame(normalised_rows, columns=headings)


# ============================================================
# HERO
# ============================================================

st.markdown(
    f"""
    <div class="hero-row">
        <div class="hero-icon">{hero_icon}</div>
        <div>
            <div class="hero-title">{transformer_title}</div>
            <div class="hero-subtitle">
                {transformer_subtitle}
            </div>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)


# ============================================================
# UPLOAD
# ============================================================

st.markdown('<div class="upload-card">', unsafe_allow_html=True)
st.markdown(f'<div class="upload-title">{upload_title}</div>', unsafe_allow_html=True)

upload_col, tips_col = st.columns([2.3, 1])

with upload_col:
    uploaded_files = st.file_uploader(
        upload_label,
        type=accepted_upload_types,
        accept_multiple_files=True,
        help=upload_help,
    )

with tips_col:
    st.markdown(
        """
        <div class="tip-box">
            <strong>💡 Tips</strong><br><br>
            ✓ Upload one file, multiple files, or a ZIP batch<br>
            ✓ Ensure files are not open or locked<br>
            ✓ Large batches may take a moment to upload
        </div>
        """,
        unsafe_allow_html=True,
    )

st.markdown("</div>", unsafe_allow_html=True)



def normalise_uploaded_files(uploaded_value):
    if uploaded_value is None:
        return []
    if isinstance(uploaded_value, list):
        return uploaded_value
    return [uploaded_value]


uploaded_file_list = normalise_uploaded_files(uploaded_files)

if uploaded_file_list:
    process_clicked = st.button(process_button_label, type="primary")
else:
    process_clicked = False
    st.info("Upload one source file, multiple source files, or a ZIP batch for the selected transformer.")


if process_clicked:
    uploaded_source_files = [
        {
            "name": f.name,
            "bytes": f.getvalue(),
        }
        for f in uploaded_file_list
    ]

    st.session_state["uploaded_source_files"] = uploaded_source_files
    st.session_state["profile_name"] = profile_name

    if active_transformer == "gas_to_ags":
        with st.spinner("Reading field gas monitoring data and building AGS..."):
            result = run_gas_from_uploaded_sources(
                uploaded_source_files,
                settings=base_settings,
                profile_name=profile_name,
            )

        st.session_state["result"] = result
        st.session_state["uploaded_zip_bytes"] = None
        st.success(
            f"Processed {len(uploaded_source_files)} field gas monitoring source file(s)."
        )
    elif active_transformer == "csv_to_ags":
        with st.spinner("Reading CSV / Excel AGS tables and building AGS..."):
            result = run_csv_ags_from_uploaded_sources(
                uploaded_source_files,
                settings=base_settings,
                profile_name=profile_name,
            )

        st.session_state["result"] = result
        st.session_state["uploaded_zip_bytes"] = None
        st.success(
            f"Processed {len(uploaded_source_files)} CSV / Excel AGS source file(s)."
        )
    else:
        if len(uploaded_file_list) != 1:
            st.error("Flowfinity field data currently expects one ZIP batch. Select one ZIP file.")
            st.stop()

        first_name = uploaded_file_list[0].name.lower()
        if not first_name.endswith(".zip"):
            st.error("Flowfinity field data currently expects a ZIP file.")
            st.stop()

        with st.spinner("Reading field data and detecting projects, locations and groups..."):
            result = run_from_zip_bytes(
                uploaded_file_list[0].getvalue(),
                settings=base_settings,
                profile_name=profile_name,
            )

        st.session_state["result"] = result
        st.session_state["uploaded_zip_bytes"] = uploaded_file_list[0].getvalue()

result = st.session_state.get("result")

if active_transformer in {"gas_to_ags", "csv_to_ags"} and result is None:
    if active_transformer == "csv_to_ags":
        st.info("Upload one CSV file, multiple CSV / Excel files, or a ZIP batch containing AGS tables.")
    else:
        st.info("Upload one Excel file, multiple Excel files, or a ZIP batch containing field gas monitoring sheets.")
    st.stop()



# ============================================================
# EMPTY STATE
# ============================================================

if result is None:
    st.markdown(
        """
        <div class="summary-grid">
            <div class="summary-card">
                <div class="summary-icon">▣</div>
                <div>
                    <div class="summary-label">Projects</div>
                    <div class="summary-value">—</div>
                    <div class="summary-caption">Waiting for upload</div>
                </div>
            </div>
            <div class="summary-card">
                <div class="summary-icon">☷</div>
                <div>
                    <div class="summary-label">Groups</div>
                    <div class="summary-value">—</div>
                    <div class="summary-caption">Waiting for upload</div>
                </div>
            </div>
            <div class="summary-card">
                <div class="summary-icon">⌖</div>
                <div>
                    <div class="summary-label">Locations</div>
                    <div class="summary-value">—</div>
                    <div class="summary-caption">Waiting for upload</div>
                </div>
            </div>
            <div class="summary-card">
                <div class="summary-icon">□</div>
                <div>
                    <div class="summary-label">Files</div>
                    <div class="summary-value">—</div>
                    <div class="summary-caption">Waiting for upload</div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.stop()


# ============================================================
# BUILD DATAFRAMES
# ============================================================

source_file_rows = [
    {
        "File": f.file_name,
        "Path": f.relative_path,
        "Dataset": f.group,
        "Project": str(f.project_id),
        "Location ID": str(f.loca_id),
        "Rows": int(f.rows),
    }
    for f in result.source_files
]

files_df = pd.DataFrame(source_file_rows)

if files_df.empty:
    st.error("No supported Flowfinity field data files were detected.")
    st.stop()

project_ids = sorted(files_df["Project"].dropna().astype(str).unique().tolist())

project_summary_df = (
    files_df.groupby("Project", dropna=False)
    .agg(
        Groups=("Dataset", "nunique"),
        Locations=("Location ID", "nunique"),
        Files=("File", "count"),
        Rows=("Rows", "sum"),
    )
    .reset_index()
    .sort_values("Project")
)

all_groups = sorted(files_df["Dataset"].dropna().astype(str).unique().tolist())

selected_project = st.selectbox(
    "Select project to inspect",
    project_ids,
    index=0,
)

project_files_df = files_df[files_df["Project"].astype(str) == str(selected_project)].copy()

if active_transformer == "gas_to_ags":
    project_file_count = project_files_df["File"].nunique()
    imported_file_count = files_df["File"].nunique()
    file_card_label = "Imports"
    file_card_caption = "Source files"
    files_metric_label = "Imports used"
else:
    project_file_count = len(project_files_df)
    imported_file_count = len(files_df)
    file_card_label = "Files"
    file_card_caption = "Imported"
    files_metric_label = "Files used"

project_location_count = project_files_df["Location ID"].nunique()
project_group_count = project_files_df["Dataset"].nunique()
project_row_count = int(project_files_df["Rows"].sum())

summary = result.summary


# ============================================================
# FULL DATA SUMMARY
# ============================================================

st.markdown(
    f"""
    <div class="summary-grid">
        <div class="summary-card">
            <div class="summary-icon">▣</div>
            <div>
                <div class="summary-label">Projects</div>
                <div class="summary-value">{len(project_ids)}</div>
                <div class="summary-caption">Detected</div>
            </div>
        </div>
        <div class="summary-card">
            <div class="summary-icon">☷</div>
            <div>
                <div class="summary-label">Groups</div>
                <div class="summary-value">{len(all_groups)}</div>
                <div class="summary-caption">Found</div>
            </div>
        </div>
        <div class="summary-card">
            <div class="summary-icon">⌖</div>
            <div>
                <div class="summary-label">Locations</div>
                <div class="summary-value">{files_df["Location ID"].nunique()}</div>
                <div class="summary-caption">Total</div>
            </div>
        </div>
        <div class="summary-card">
            <div class="summary-icon">□</div>
            <div>
                <div class="summary-label">Files</div>
                <div class="summary-value">{len(files_df)}</div>
                <div class="summary-caption">Imported</div>
            </div>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)


# ============================================================
# PROJECTS + GROUP CHIPS
# ============================================================

st.markdown('<div class="section-card">', unsafe_allow_html=True)
st.markdown('<div class="section-heading">Projects detected</div>', unsafe_allow_html=True)

display_project_summary = project_summary_df.copy()
display_project_summary["Status"] = "Ready"
st.dataframe(display_project_summary, use_container_width=True, hide_index=True)

st.markdown("</div>", unsafe_allow_html=True)


st.markdown('<div class="section-card">', unsafe_allow_html=True)
st.markdown('<div class="section-heading">Groups detected</div>', unsafe_allow_html=True)

chips = "".join(
    f'<div class="group-chip">{html.escape(group)}</div>'
    for group in all_groups
)

st.markdown(f'<div class="chip-row">{chips}</div>', unsafe_allow_html=True)
st.markdown("</div>", unsafe_allow_html=True)


# ============================================================
# SELECTED PROJECT REVIEW
# ============================================================

st.subheader(f"Project review: {selected_project}")
st.caption("Project metadata is optional. If left unchanged, the app uses the project and location values detected from the uploaded source data.")
st.caption("Project metadata is optional. If left unchanged, the app uses the project and location values detected from the uploaded source data.")

col1, col2, col3, col4, col5, col6 = st.columns(6)
col1.metric("AGS version", summary["ags_version"])
col2.metric(files_metric_label, project_file_count)
col3.metric("Locations", project_location_count)
col4.metric("Groups", project_group_count)
col5.metric("Rows", project_row_count)
col6.metric("Errors", summary["errors"])


# ============================================================
# METADATA
# ============================================================

with st.expander("Optional project metadata override", expanded=False):
    metadata_df = pd.DataFrame(
        [
            {
                "Project ID": selected_project,
                "Project name": selected_project,
                "Project location": "",
                "Client": "",
                "Contractor": "",
                "Engineer": "",
                "Project notes": "",
            }
        ]
    )

    edited_metadata_df = st.data_editor(
        metadata_df,
        use_container_width=True,
        hide_index=True,
        num_rows="fixed",
        key=f"project_metadata_editor_{selected_project}",
    )

    selected_project_metadata = edited_metadata_df.iloc[0].to_dict() if not edited_metadata_df.empty else {}

export_settings = {
    **base_settings,
    "project_id": selected_project,
    "project_name": selected_project,
    "project_location": "",
    "project_client": "",
    "project_contractor": "",
    "project_engineer": "",
    "project_memo": "",
}

if "selected_project_metadata" in locals():
    export_settings.update(
        {
            "project_id": selected_project_metadata.get("Project ID", selected_project),
            "project_name": selected_project_metadata.get("Project name", selected_project),
            "project_location": selected_project_metadata.get("Project location", ""),
            "project_client": selected_project_metadata.get("Client", ""),
            "project_contractor": selected_project_metadata.get("Contractor", ""),
            "project_engineer": selected_project_metadata.get("Engineer", ""),
            "project_memo": selected_project_metadata.get("Project notes", ""),
        }
    )


rebuild_clicked = st.button("Rebuild AGS with selected project metadata")

if rebuild_clicked:
    if active_transformer == "gas_to_ags":
        uploaded_source_files = st.session_state.get("uploaded_source_files", [])

        if uploaded_source_files:
            with st.spinner("Rebuilding gas monitoring AGS with selected project metadata..."):
                result = run_gas_from_uploaded_sources(
                    uploaded_source_files,
                    settings=export_settings,
                    profile_name=profile_name,
                )

            st.session_state["result"] = result
            st.success("Gas monitoring AGS rebuilt with selected project metadata.")
            st.rerun()
    elif active_transformer == "csv_to_ags":
        uploaded_source_files = st.session_state.get("uploaded_source_files", [])

        if uploaded_source_files:
            with st.spinner("Rebuilding CSV / Excel AGS with selected project metadata..."):
                result = run_csv_ags_from_uploaded_sources(
                    uploaded_source_files,
                    settings=export_settings,
                    profile_name=profile_name,
                )

            st.session_state["result"] = result
            st.success("CSV / Excel AGS rebuilt with selected project metadata.")
            st.rerun()
    else:
        zip_bytes = st.session_state.get("uploaded_zip_bytes")

        if zip_bytes:
            with st.spinner("Rebuilding AGS with selected project metadata..."):
                result = run_from_zip_bytes(
                    zip_bytes,
                    settings=export_settings,
                    profile_name=profile_name,
                )

            st.session_state["result"] = result
            st.success("AGS rebuilt with selected project metadata.")
            st.rerun()


# ============================================================
# TABS
# ============================================================

review_tabs = ["Datasets", "Locations", "Source files", "Issues", "Group preview", "Export"]

active_review_tab = st.radio(
    "Review section",
    review_tabs,
    horizontal=True,
    key="active_review_tab",
    label_visibility="collapsed",
)


if active_review_tab == "Datasets":
    st.subheader(f"Datasets for project {selected_project}")

    group_df = (
        project_files_df.groupby("Dataset", dropna=False)
        .agg(
            Files=("File", "count"),
            Locations=("Location ID", "nunique"),
            Rows=("Rows", "sum"),
        )
        .reset_index()
        .sort_values("Dataset")
    )

    st.dataframe(group_df, use_container_width=True, hide_index=True)


if active_review_tab == "Locations":
    st.subheader(f"Locations for project {selected_project}")

    location_df = (
        project_files_df.groupby("Location ID", dropna=False)
        .agg(
            Files=("File", "count"),
            Groups=("Dataset", "nunique"),
            Rows=("Rows", "sum"),
        )
        .reset_index()
        .sort_values("Location ID")
    )

    st.dataframe(location_df, use_container_width=True, hide_index=True)

    selected_location_options = location_df["Location ID"].dropna().astype(str).tolist()

    if selected_location_options:
        selected_location = st.selectbox(
            "Select location to inspect",
            selected_location_options,
            index=0,
        )

        location_files_df = project_files_df[
            project_files_df["Location ID"].astype(str) == str(selected_location)
        ]

        st.markdown(f"### Location review: `{selected_location}`")

        location_group_df = (
            location_files_df.groupby("Dataset", dropna=False)
            .agg(
                Files=("File", "count"),
                Rows=("Rows", "sum"),
            )
            .reset_index()
            .sort_values("Dataset")
        )

        st.dataframe(location_group_df, use_container_width=True, hide_index=True)


if active_review_tab == "Source files":
    st.subheader(f"Source files for project {selected_project}" if active_transformer != "gas_to_ags" else f"Source records for project {selected_project}")

    st.dataframe(
        project_files_df.sort_values(["Location ID", "Dataset", "File"]),
        use_container_width=True,
        hide_index=True,
    )

    if result.skipped_files:
        with st.expander("Skipped files from full package"):
            for skipped in result.skipped_files:
                st.write(skipped)


if active_review_tab == "Issues":
    st.subheader(f"Issues for project {selected_project}")

    issue_rows = [
        {
            "Severity": issue.severity,
            "Dataset": issue.group,
            "Location": issue.location,
            "File": issue.file,
            "Message": issue.message,
        }
        for issue in result.issues
    ]

    issues_df = pd.DataFrame(issue_rows)

    if issues_df.empty:
        st.success("No processing issues detected.")
    else:
        project_loca_ids = set(project_files_df["Location ID"].astype(str).tolist())
        project_issue_df = issues_df[
            issues_df["Location"].astype(str).isin(project_loca_ids)
        ]

        if project_issue_df.empty:
            st.success("No processing issues detected for this project.")
        else:
            st.dataframe(project_issue_df, use_container_width=True, hide_index=True)




if active_review_tab == "Group preview":
    st.subheader("Preview individual AGS group")

    group_tables = parse_ags_text_to_group_tables(result.ags_text)
    group_names = list(group_tables.keys())

    if not group_names:
        st.info("No AGS groups available to preview.")
    else:
        selected_group_name = st.selectbox(
            "Select AGS group to preview",
            group_names,
            key="selected_group_preview_name",
        )

        selected_table = group_tables[selected_group_name]
        selected_headings = selected_table.get("headings", [])
        selected_rows = selected_table.get("rows", [])

        st.caption("Selected group")
        st.code(selected_group_name, language="text")

        preview_mode = st.radio(
            "Preview format",
            ["AGS", "CSV table", "JSON"],
            horizontal=True,
            key="group_preview_format",
        )

        c1, c2, c3 = st.columns(3)
        c1.metric("Rows", len(selected_rows))
        c2.metric("Fields", len(selected_headings))
        c3.metric("Format", preview_mode)

        if preview_mode == "AGS":
            group_text = extract_ags_group_text(result.ags_text, selected_group_name)
            st.text_area(
                "AGS group text",
                value=group_text,
                height=420,
                key=f"ags_group_text_{selected_group_name}",
            )

            st.download_button(
                "Download selected group as AGS text",
                data=group_text.encode("utf-8"),
                file_name=f"{safe_export_name(selected_group_name)}.ags.txt",
                mime="text/plain",
                key=f"download_group_ags_{selected_group_name}",
            )

        elif preview_mode == "CSV table":
            if selected_headings:
                preview_df = pd.DataFrame(
                    selected_rows,
                    columns=unique_table_fields(selected_headings),
                )
                st.dataframe(preview_df, use_container_width=True, hide_index=True)
            else:
                st.info("This group has no headings to display as a table.")

            st.download_button(
                "Download selected group CSV",
                data=ags_group_to_csv_bytes(selected_table),
                file_name=f"{safe_export_name(selected_group_name)}.csv",
                mime="text/csv",
                key=f"download_group_csv_{selected_group_name}",
            )

        else:
            st.json(
                ags_group_to_json_payload(selected_group_name, selected_table),
                expanded=False,
            )

            st.download_button(
                "Download selected group JSON",
                data=ags_group_to_json_bytes(selected_group_name, selected_table),
                file_name=f"{safe_export_name(selected_group_name)}.json",
                mime="application/json",
                key=f"download_group_json_{selected_group_name}",
            )


if active_review_tab == "Export":
    st.subheader("Export")

    export_project_id = (
        st.session_state.get("selected_project_id")
        or next(iter(result.project_counts.keys()), "")
        or "ags_export"
    )
    export_base_name = safe_export_name(export_project_id, "ags_export")

    group_tables = parse_ags_text_to_group_tables(result.ags_text)

    st.caption(
        f"Export package contains {len(group_tables)} AGS group(s). "
        "CSV and JSON are generated from the same AGS output shown in preview."
    )

    col_ags, col_csv, col_json, col_excel, col_full = st.columns(5)

    with col_ags:
        st.download_button(
            "Download AGS",
            data=result.ags_text.encode("utf-8"),
            file_name=f"{export_base_name}.ags",
            mime="text/plain",
            use_container_width=True,
            key="download_full_ags",
        )

    with col_csv:
        st.download_button(
            "Download CSV ZIP",
            data=build_table_package_zip(
                result.ags_text,
                f"{export_base_name}_csv_tables",
                include_csv=True,
            ),
            file_name=f"{export_base_name}_csv_tables.zip",
            mime="application/zip",
            use_container_width=True,
            key="download_csv_package",
        )

    with col_json:
        st.download_button(
            "Download JSON ZIP",
            data=build_table_package_zip(
                result.ags_text,
                f"{export_base_name}_json_tables",
                include_json=True,
            ),
            file_name=f"{export_base_name}_json_tables.zip",
            mime="application/zip",
            use_container_width=True,
            key="download_json_package",
        )

    with col_excel:
        st.download_button(
            "Download Excel",
            data=build_excel_workbook_bytes(
                result.ags_text,
                f"{export_base_name}_tables",
            ),
            file_name=f"{export_base_name}_tables.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="download_excel_workbook",
        )

    with col_full:
        st.download_button(
            "Download full package ZIP",
            data=build_table_package_zip(
                result.ags_text,
                f"{export_base_name}_full_package",
                include_ags=True,
                include_csv=True,
                include_json=True,
                include_excel=True,
            ),
            file_name=f"{export_base_name}_full_package.zip",
            mime="application/zip",
            use_container_width=True,
            key="download_full_package",
        )

    with st.expander("Package structure", expanded=False):
        st.code(
            """full_package.zip
├─ manifest.json
├─ <project>.ags
├─ <project>.xlsx
├─ csv/
│  ├─ PROJ.csv
│  ├─ LOCA.csv
│  └─ ...
└─ json/
   ├─ PROJ.json
   ├─ LOCA.json
   └─ ...""",
            language="text",
        )

